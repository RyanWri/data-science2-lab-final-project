from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from src_code.dictionary import HebEngTranslator

class ExceltUtils:
  
  @classmethod
  def adjust_columns_width(cls, original: Worksheet, target: Worksheet) -> None:
    for col in original.columns:
      max_length = 0
      for cell in col:
        try:
          if len(str(cell.value)) > max_length:
              max_length = len(cell.value)
        except:
          pass
      adjusted_width = (max_length + 10)
      target.column_dimensions[col[0].column_letter].width = adjusted_width

  @classmethod
  def adjust_cells_font_and_font_size(cls, target: Worksheet) -> None:
    for row in target.iter_rows():
      for cell in row:
        cell.font = Font(name="Arial", size=13)
  
  @classmethod
  def extract_sheet(cls, document: Workbook, sheet_name: str) -> Workbook:
    result = Workbook()
    result.remove(result.active)
    target_sheet = result.create_sheet(title=sheet_name)
    cls.adjust_columns_width(document[sheet_name], target_sheet)
    for row in document[sheet_name].iter_rows():
      for cell in row:
        copied_cell = target_sheet.cell(row=cell.row, column=cell.col_idx, value=cell.value)
        copied_cell.number_format = cell.number_format
    cls.adjust_cells_font_and_font_size(document[sheet_name])
    return result
  
  @classmethod
  def read_file(cls, abs_file_path: str) -> Workbook:
    return load_workbook(abs_file_path)
  
  @classmethod
  def save_to_file(cls, document: Workbook, abs_file_path: str) -> None:
    document.save(abs_file_path)

  @classmethod
  def translate_cell_values_to_english(cls, document: Workbook, sheet_name: str) -> None:
    for row in document[sheet_name].iter_rows():
      for cell in row:
        if cell.value is not None and type(cell.value) is str:
          cell.value = HebEngTranslator.to_eng(cell.value)

  @classmethod
  def translate_cell_values_to_hebrew(cls, document: Workbook, sheet_name: str) -> None:
    for row in document[sheet_name].iter_rows():
      for cell in row:
        if cell.value is not None and type(cell.value) is str:
          cell.value = HebEngTranslator.to_heb(cell.value)
  
  @classmethod
  def validate_ascii_encoded(cls, document: Workbook, sheet_name: str) -> None:
    for row in document[sheet_name].iter_rows():
      for cell in row:
        try:
          if type(cell.value) is str:
            cell.value.encode("ascii")
        except:
          raise RuntimeError(f"{cell.value} is not ASCII encoded")
